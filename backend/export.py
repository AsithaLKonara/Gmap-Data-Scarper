from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from models import User, Job, Lead, Plan
from database import get_db
from auth import get_current_user
import logging
import json
import csv
import io
import zipfile
from datetime import datetime, timedelta
from enum import Enum
import asyncio
import aiofiles
import os
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

logger = logging.getLogger("export")

router = APIRouter(prefix="/api/export", tags=["export"])

class ExportFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"
    PDF = "pdf"

class ExportType(str, Enum):
    JOBS = "jobs"
    LEADS = "leads"
    ANALYTICS = "analytics"
    ALL = "all"

class ExportRequest(BaseModel):
    export_type: ExportType
    format: ExportFormat
    filters: Optional[Dict[str, Any]] = None
    include_metadata: bool = True
    compress: bool = False

class ExportStatus(BaseModel):
    id: str
    status: str  # pending, processing, completed, failed
    progress: int
    total_items: int
    processed_items: int
    download_url: Optional[str] = None
    error_message: Optional[str] = None
    created_at: datetime
    completed_at: Optional[datetime] = None

    class Config:
        from_attributes = True

# In-memory export status (in production, use database)
export_statuses: Dict[str, ExportStatus] = {}

class DataExporter:
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    async def export_jobs(self, user_id: int, format: ExportFormat, filters: Dict = None) -> str:
        """Export jobs data"""
        try:
            # Get jobs from database
            db = next(get_db())
            query = db.query(Job).filter(Job.user_id == user_id)
            
            if filters:
                if filters.get('status'):
                    query = query.filter(Job.status == filters['status'])
                if filters.get('date_from'):
                    query = query.filter(Job.created_at >= filters['date_from'])
                if filters.get('date_to'):
                    query = query.filter(Job.created_at <= filters['date_to'])
            
            jobs = query.all()
            
            # Convert to export format
            if format == ExportFormat.CSV:
                return await self._export_to_csv(jobs, "jobs")
            elif format == ExportFormat.JSON:
                return await self._export_to_json(jobs, "jobs")
            elif format == ExportFormat.EXCEL:
                return await self._export_to_excel(jobs, "jobs")
            elif format == ExportFormat.PDF:
                return await self._export_to_pdf(jobs, "jobs")
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.exception("Error exporting jobs")
            raise
    
    async def export_leads(self, user_id: int, format: ExportFormat, filters: Dict = None) -> str:
        """Export leads data"""
        try:
            # Get leads from database
            db = next(get_db())
            query = db.query(Lead).filter(Lead.user_id == user_id)
            
            if filters:
                if filters.get('status'):
                    query = query.filter(Lead.status == filters['status'])
                if filters.get('tag'):
                    query = query.filter(Lead.tag == filters['tag'])
                if filters.get('date_from'):
                    query = query.filter(Lead.created_at >= filters['date_from'])
                if filters.get('date_to'):
                    query = query.filter(Lead.created_at <= filters['date_to'])
            
            leads = query.all()
            
            # Convert to export format
            if format == ExportFormat.CSV:
                return await self._export_to_csv(leads, "leads")
            elif format == ExportFormat.JSON:
                return await self._export_to_json(leads, "leads")
            elif format == ExportFormat.EXCEL:
                return await self._export_to_excel(leads, "leads")
            elif format == ExportFormat.PDF:
                return await self._export_to_pdf(leads, "leads")
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.exception("Error exporting leads")
            raise
    
    async def export_analytics(self, user_id: int, format: ExportFormat, filters: Dict = None) -> str:
        """Export analytics data"""
        try:
            # Generate analytics data
            db = next(get_db())
            
            # Job statistics
            total_jobs = db.query(Job).filter(Job.user_id == user_id).count()
            completed_jobs = db.query(Job).filter(Job.user_id == user_id, Job.status == 'completed').count()
            failed_jobs = db.query(Job).filter(Job.user_id == user_id, Job.status == 'failed').count()
            
            # Lead statistics
            total_leads = db.query(Lead).filter(Lead.user_id == user_id).count()
            new_leads = db.query(Lead).filter(Lead.user_id == user_id, Lead.status == 'new').count()
            converted_leads = db.query(Lead).filter(Lead.user_id == user_id, Lead.status == 'converted').count()
            
            analytics_data = {
                'job_statistics': {
                    'total_jobs': total_jobs,
                    'completed_jobs': completed_jobs,
                    'failed_jobs': failed_jobs,
                    'success_rate': (completed_jobs / total_jobs * 100) if total_jobs > 0 else 0
                },
                'lead_statistics': {
                    'total_leads': total_leads,
                    'new_leads': new_leads,
                    'converted_leads': converted_leads,
                    'conversion_rate': (converted_leads / total_leads * 100) if total_leads > 0 else 0
                },
                'export_date': datetime.utcnow().isoformat(),
                'user_id': user_id
            }
            
            # Convert to export format
            if format == ExportFormat.JSON:
                return await self._export_to_json([analytics_data], "analytics")
            elif format == ExportFormat.CSV:
                return await self._export_to_csv([analytics_data], "analytics")
            elif format == ExportFormat.EXCEL:
                return await self._export_to_excel([analytics_data], "analytics")
            elif format == ExportFormat.PDF:
                return await self._export_to_pdf([analytics_data], "analytics")
            else:
                raise ValueError(f"Unsupported format for analytics: {format}")
                
        except Exception as e:
            logger.exception("Error exporting analytics")
            raise
    
    async def _export_to_csv(self, data: List, data_type: str) -> str:
        """Export data to CSV format"""
        try:
            if not data:
                raise ValueError("No data to export")
            
            # Generate filename
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"{data_type}_{timestamp}.csv"
            filepath = os.path.join(self.export_dir, filename)
            
            # Write CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                if data_type == "jobs":
                    fieldnames = ['id', 'queries', 'status', 'created_at', 'updated_at']
                elif data_type == "leads":
                    fieldnames = ['id', 'name', 'email', 'phone', 'company', 'tag', 'status', 'created_at']
                else:
                    fieldnames = list(data[0].keys()) if data else []
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for item in data:
                    if hasattr(item, '__dict__'):
                        row = {k: v for k, v in item.__dict__.items() if k in fieldnames}
                    else:
                        row = item
                    writer.writerow(row)
            
            return filename
            
        except Exception as e:
            logger.exception("Error exporting to CSV")
            raise
    
    async def _export_to_json(self, data: List, data_type: str) -> str:
        """Export data to JSON format"""
        try:
            if not data:
                raise ValueError("No data to export")
            
            # Generate filename
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"{data_type}_{timestamp}.json"
            filepath = os.path.join(self.export_dir, filename)
            
            # Convert to serializable format
            serializable_data = []
            for item in data:
                if hasattr(item, '__dict__'):
                    item_dict = {k: v for k, v in item.__dict__.items() if not k.startswith('_')}
                    # Convert datetime objects
                    for key, value in item_dict.items():
                        if isinstance(value, datetime):
                            item_dict[key] = value.isoformat()
                    serializable_data.append(item_dict)
                else:
                    serializable_data.append(item)
            
            # Write JSON
            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(serializable_data, jsonfile, indent=2, ensure_ascii=False)
            
            return filename
            
        except Exception as e:
            logger.exception("Error exporting to JSON")
            raise
    
    async def _export_to_excel(self, data: List, data_type: str) -> str:
        """Export data to Excel (XLSX) format"""
        try:
            if not data:
                raise ValueError("No data to export")
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"{data_type}_{timestamp}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            wb = openpyxl.Workbook()
            ws = wb.active
            # Determine fields
            if data_type == "jobs":
                fieldnames = ['id', 'queries', 'status', 'created_at', 'updated_at']
            elif data_type == "leads":
                fieldnames = ['id', 'name', 'email', 'phone', 'company', 'tag', 'status', 'created_at']
            else:
                fieldnames = list(data[0].keys()) if data else []
            ws.append(fieldnames)
            for item in data:
                if hasattr(item, '__dict__'):
                    row = [getattr(item, k, '') for k in fieldnames]
                else:
                    row = [item.get(k, '') for k in fieldnames]
                ws.append(row)
            for i, col in enumerate(fieldnames, 1):
                ws.column_dimensions[get_column_letter(i)].width = 18
            wb.save(filepath)
            return filename
        except Exception as e:
            logger.exception("Error exporting to Excel")
            raise

    async def _export_to_pdf(self, data: List, data_type: str) -> str:
        """Export data to PDF format"""
        try:
            if not data:
                raise ValueError("No data to export")
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"{data_type}_{timestamp}.pdf"
            filepath = os.path.join(self.export_dir, filename)
            c = canvas.Canvas(filepath, pagesize=letter)
            width, height = letter
            # Determine fields
            if data_type == "jobs":
                fieldnames = ['id', 'queries', 'status', 'created_at', 'updated_at']
            elif data_type == "leads":
                fieldnames = ['id', 'name', 'email', 'phone', 'company', 'tag', 'status', 'created_at']
            else:
                fieldnames = list(data[0].keys()) if data else []
            y = height - 40
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, f"{data_type.capitalize()} Export - {timestamp}")
            y -= 30
            c.setFont("Helvetica-Bold", 10)
            for i, col in enumerate(fieldnames):
                c.drawString(40 + i*100, y, str(col))
            y -= 20
            c.setFont("Helvetica", 9)
            for item in data:
                if y < 60:
                    c.showPage()
                    y = height - 40
                if hasattr(item, '__dict__'):
                    row = [getattr(item, k, '') for k in fieldnames]
                else:
                    row = [item.get(k, '') for k in fieldnames]
                for i, val in enumerate(row):
                    c.drawString(40 + i*100, y, str(val))
                y -= 18
            c.save()
            return filename
        except Exception as e:
            logger.exception("Error exporting to PDF")
            raise

# Initialize exporter
exporter = DataExporter()

def get_allowed_export_formats(user: User, db: Session):
    plan = db.query(Plan).filter(Plan.name == user.plan).first()
    if not plan or not plan.limits:
        return ["csv"]  # Default to CSV only if plan not found
    try:
        limits = json.loads(plan.limits)
        return limits.get("export_formats", ["csv"])
    except Exception:
        return ["csv"]

@router.post("/start", response_model=ExportStatus)
async def start_export(
    export_request: ExportRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    try:
        # Enforce allowed export formats by plan
        allowed_formats = get_allowed_export_formats(user, db)
        requested_format = export_request.format.value if hasattr(export_request.format, 'value') else str(export_request.format)
        if requested_format not in allowed_formats:
            raise HTTPException(status_code=403, detail=f"Your plan does not allow exporting in {requested_format.upper()} format. Allowed: {', '.join(allowed_formats).upper()}")
        
        # Generate export ID
        export_id = f"export_{user.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        
        # Create export status
        export_status = ExportStatus(
            id=export_id,
            status="pending",
            progress=0,
            total_items=0,
            processed_items=0,
            download_url=None,
            error_message=None,
            created_at=datetime.utcnow(),
            completed_at=None
        )
        
        export_statuses[export_id] = export_status
        
        # Start background export
        background_tasks.add_task(
            process_export,
            export_id,
            export_request,
            user.id
        )
        
        logger.info(f"Export started: {export_id} by user {user.id}")
        return export_status
        
    except Exception as e:
        logger.exception("Error starting export")
        raise HTTPException(status_code=500, detail="Failed to start export")

async def process_export(export_id: str, export_request: ExportRequest, user_id: int):
    """Background task to process export"""
    try:
        export_status = export_statuses[export_id]
        export_status.status = "processing"
        export_status.progress = 10
        
        # Process export based on type
        if export_request.export_type == ExportType.JOBS:
            filename = await exporter.export_jobs(user_id, export_request.format, export_request.filters)
        elif export_request.export_type == ExportType.LEADS:
            filename = await exporter.export_leads(user_id, export_request.format, export_request.filters)
        elif export_request.export_type == ExportType.ANALYTICS:
            filename = await exporter.export_analytics(user_id, export_request.format, export_request.filters)
        elif export_request.export_type == ExportType.ALL:
            # Export all data types
            jobs_file = await exporter.export_jobs(user_id, export_request.format, export_request.filters)
            leads_file = await exporter.export_leads(user_id, export_request.format, export_request.filters)
            analytics_file = await exporter.export_analytics(user_id, export_request.format, export_request.filters)
            
            # Create zip file
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"all_data_{timestamp}.zip"
            zip_path = os.path.join(exporter.export_dir, filename)
            
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for file in [jobs_file, leads_file, analytics_file]:
                    file_path = os.path.join(exporter.export_dir, file)
                    if os.path.exists(file_path):
                        zipf.write(file_path, file)
        else:
            raise ValueError(f"Unsupported export type: {export_request.export_type}")
        
        # Update status
        export_status.status = "completed"
        export_status.progress = 100
        export_status.download_url = f"/api/export/download/{filename}"
        export_status.completed_at = datetime.utcnow()
        
        logger.info(f"Export completed: {export_id}")
        
    except Exception as e:
        logger.exception(f"Error processing export {export_id}")
        export_status = export_statuses[export_id]
        export_status.status = "failed"
        export_status.error_message = str(e)
        export_status.completed_at = datetime.utcnow()

@router.get("/status/{export_id}", response_model=ExportStatus)
def get_export_status(export_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        export_status = export_statuses.get(export_id)
        if not export_status:
            raise HTTPException(status_code=404, detail="Export not found")
        
        return export_status
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error getting export status")
        raise HTTPException(status_code=500, detail="Failed to get export status")

@router.get("/download/{filename}")
def download_export(filename: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        file_path = os.path.join(exporter.export_dir, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Export file not found")
        
        # Return file for download
        from fastapi.responses import FileResponse
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/octet-stream'
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error downloading export")
        raise HTTPException(status_code=500, detail="Failed to download export")

@router.get("/history", response_model=List[ExportStatus])
def get_export_history(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        # Return user's export history
        user_exports = [
            status for status in export_statuses.values()
            if status.id.startswith(f"export_{user.id}_")
        ]
        return sorted(user_exports, key=lambda x: x.created_at, reverse=True)
    except Exception as e:
        logger.exception("Error getting export history")
        raise HTTPException(status_code=500, detail="Failed to get export history") 